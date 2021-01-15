#-------------------------------------------------------------------------------
# Name:        ora_excel_read.py
# Purpose:     a collection of reusable functions
# Author:      Mike Martin
# Created:     26/12/2019
# Licence:     <your licence>
# Definitions:
#   spin_up
#
# Description:
#
#
#-------------------------------------------------------------------------------
#!/usr/bin/env python

__prog__ = 'ora_excel_read.py'
__version__ = '0.0.0'

# Version history
# ---------------
#
import os
from glob import glob
from calendar import monthrange
from openpyxl import load_workbook
from pandas import Series, read_excel
from zipfile import BadZipFile

from ora_excel_write import retrieve_output_xls_files
from ora_water_model import add_pet_to_weather
from ora_cn_fns import plant_inputs_crops_distribution
from ora_low_level_fns import average_weather

METRIC_LIST = list(['precip', 'tair'])
MNTH_NAMES_SHORT = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
REQUIRED_SHEET_NAMES = list(['Inputs1- Farm location','N constants', 'Crop parms','Org Waste parms','Weather'])

ERR_MESS_SHEET = '*** Error *** reading sheet '

def check_excel_input_file(form, xls_inp_fname):
    '''
    validate selected Excel file
    '''
    fileOkFlag = True

    if not os.path.isfile(xls_inp_fname):
        return None

    form.w_study.setText('Study not set')
    form.settings['inp_dir'] = ''
    form.w_soil_cn.setEnabled(False)
    form.w_disp_out.setEnabled(False)

    print('Loading parameters and weather inputs file: ' + xls_inp_fname)
    try:
        wb_obj = load_workbook(xls_inp_fname, data_only = True)
        sheet_names = wb_obj.sheetnames
    except (PermissionError, BadZipFile) as e:
        print('Error: ' + str(e))
        return None

    wb_obj.close()

    # all required sheets must be present
    # ===================================
    for sheet in REQUIRED_SHEET_NAMES:
        if sheet not in sheet_names:
            fileOkFlag = False
            break

    if fileOkFlag:
        mess = '\tinputs file is valid'
        form.w_soil_cn.setEnabled(True)

        study, latitude, longitude = _read_location_sheet(xls_inp_fname, 'Inputs1- Farm location', 13)
        study_desc = 'Study: ' + study
        study_desc += '\tLatitude: {}'.format(latitude)
        study_desc += '\tLongitude: {}'.format(longitude)
        form.w_study.setText(study_desc)

        form.settings['study'] = study
        form.settings['inp_dir'], dummy = os.path.split(xls_inp_fname)
        retrieve_output_xls_files(form, study)

    else:
        mess = 'Required sheet ' + sheet + ' is not present - please check file'

    print(mess + '\n')
    return mess

def _read_location_sheet(xls_fname, sheet_name, skip_until):
    '''

    '''
    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(2,4))

    location = data.dropna(how='all')
    location = location.rename(columns={'Unnamed: 2': 'Description', 'Unnamed: 3': 'Value'})

    study = location['Value'].values[1]
    latitude  = location['Value'].values[2]
    longitude = location['Value'].values[3]

    return study, latitude, longitude

def _read_n_constants_sheet(xls_fname, sheet_name, skip_until):
    '''
    r_dry is an environmental constant
    '''

    n_parm_names = list(['atmos_n_depos', 'prop_atmos_dep_no3', 'no3_min', 'k_nitrif',
                      'n_denit_max', 'n_d50', 'prop_n2o_fc', 'prop_nitrif_gas', 'prop_nitrif_no',
                      'precip_critic', 'prop_volat', 'prop_atmos_dep_nh4', 'c_n_rat_som', 'r_dry', 'k_c_rate'])

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(1,3))
    n_parms_df = data.dropna(how='all')
    n_parms = {}
    for indx, defn in enumerate(n_parm_names):
        n_parms[defn]  = n_parms_df['Value'].values[indx]

    return n_parms

def _extract_weather(pettmp_dframe, year_strt, year_end, indx):

    pettemp = []
    for yr in range(year_strt, year_end):
        slice = list(pettmp_dframe[yr][indx:indx + 12])
        pettemp += slice

    return pettemp

def _read_weather_sheet(xls_fname, sheet_name, skip_until):
    '''
    reads weather
    '''
    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until), usecols=range(3, 16))
    pettmp_dframe = data.dropna(how='all')
    year_strt = pettmp_dframe.columns[2]
    ncols = len(pettmp_dframe.columns)
    nyears = ncols - 3
    year_end = year_strt + nyears
    col_names = ['Month', 'Season'] + [year_strt + val for val in range(nyears)] + ['dummy']
    pettmp_dframe.columns = col_names

    # reorganise into total rainfall (mm month-1) and average air temperature (°C) for steady state and forward runs
    # ==============================================================================================================
    pettmp_ss = {}
    for indx, metric in zip(list([0, 14]), METRIC_LIST):
        pettmp_ss[metric] = _extract_weather(pettmp_dframe, year_strt, year_end, indx)

    pettmp_fwd = {}
    for indx, metric in zip(list([28, 42]), METRIC_LIST):
        pettmp_fwd[metric] = _extract_weather(pettmp_dframe, year_strt, year_end, indx)

    return pettmp_ss, pettmp_fwd

def _read_crop_vars(xls_fname, sheet_name):
    '''
    read maximum rooting depths etc. for each crop from sheet A1c
    '''
    crop_parm_names = Series(list(['lu_code', 'rat_dpm_rpm', 'harv_indx', 'prop_npp_to_pi', 'max_root_dpth',
                    'max_root_dpth_orig', 'sow_mnth', 'harv_mnth', 'max_yld', 'dummy1',
                    'c_n_rat_pi', 'n_sply_min', 'n_sply_opt', 'n_respns_coef', 'fert_use_eff',
                    'dummy3', 'n_rcoef_a', 'n_rcoef_b', 'n_rcoef_c', 'n_rcoef_d', 'gdds_scle_factr','iws_scle_factr']))

    data = read_excel(xls_fname, sheet_name)
    data = data.dropna(how='all')
    try:
        crop_dframe = data.set_index(crop_parm_names)
        crop_vars = crop_dframe.to_dict()
    except ValueError as err:

        print(ERR_MESS_SHEET + sheet_name + ' ' + str(err))
        return None

    # discard unwanted entries
    # ========================
    for crop in ['Crop', 'None', 'Null']:
        del(crop_vars[crop])

    for crop_name in crop_vars:

        # clean data
        # ==========
        for var in ['harv_mnth', 'sow_mnth', 'lu_code']:
            crop_vars[crop_name][var] = int(crop_vars[crop_name][var])

        # number of months in the growing season
        # ======================================
        harv_mnth = crop_vars[crop_name]['harv_mnth']
        sow_mnth = crop_vars[crop_name]['sow_mnth']
        if sow_mnth > harv_mnth:
            harv_mnth += 12
        t_grow = harv_mnth - sow_mnth + 1
        crop_vars[crop_name]['pi_tonnes'], crop_vars[crop_name]['pi_prop'] = plant_inputs_crops_distribution(t_grow)
        crop_vars[crop_name]['t_grow'] = t_grow

    return crop_vars

def _read_organic_waste_sheet(xls_fname, sheet_name, skip_until):
    '''
    read Organic waste parameters
    added  - see (eq.2.1.12) and (eq.2.1.13)
    TODO percentages are converted to fraction
    '''
    ow_parms_names = Series(list(['c_n_rat', 'prop_nh4', 'rat_dpm_hum_ow', 'prop_iom_ow', 'pcnt_c', 'min_e_pcnt_wd',
                                  'max_e_pcnt_wd', 'ann_c_input', 'pcnt_urea']))

    data = read_excel(xls_fname, sheet_name, skiprows=range(0, skip_until))
    data = data.dropna(how='all')
    try:
        ow_dframe = data.set_index(ow_parms_names)
        all_ow_parms = ow_dframe.to_dict()
    except ValueError as err:
        print(ERR_MESS_SHEET + sheet_name + ' ' + str(err))
        all_ow_parms = None

    return all_ow_parms

class ReadStudy(object, ):

    def __init__(self, xls_inp_fname, out_dir, output_excel = True):
        '''
        read location sheet from ORATOR inputs Excel file
        '''

        self.output_excel = output_excel
        self.out_dir = out_dir

        # Farm location
        # =============
        self.study_name, self.latitude, self.longitude \
                                        = _read_location_sheet(xls_inp_fname, 'Inputs1- Farm location', 13)

def _add_tgdd_to_weather(tair_list):
    '''
    growing degree days indicates the cumulative temperature when plant growth is assumed to be possible (above 5°C)
    '''

    imnth = 1
    grow_dds = []
    for tair in tair_list:

        dummy, ndays = monthrange(2011, imnth)
        n_grow_days = max(0, ndays * (tair - 5))    #  (eq.3.2.2)
        grow_dds.append(round(n_grow_days,3))

        imnth += 1
        if imnth > 12:
            imnth = 1

    return grow_dds

class ReadWeather(object, ):

    def __init__(self, xls_inp_fname, latitude):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading weather sheet...')

        pettmp_ss, pettmp_fwd = _read_weather_sheet(xls_inp_fname, 'Weather', 14)

        # generate PET from weather
        # =========================
        self.pettmp_ss = add_pet_to_weather(latitude, pettmp_ss)
        self.pettmp_fwd = add_pet_to_weather(latitude, pettmp_fwd)

        # growing degree days
        # ==================
        self.pettmp_ss['grow_dds'] = _add_tgdd_to_weather(pettmp_ss['tair'])
        self.pettmp_fwd['grow_dds'] = _add_tgdd_to_weather(pettmp_fwd['tair'])

        # average monthly precip and temp is required for spin up
        # =======================================================
        self.ave_precip_ss, self.ave_temp_ss, self.ave_pet_ss = \
                            average_weather(latitude, self.pettmp_ss['precip'], self.pettmp_ss['tair'])

        # get average annual rain and temperature of first 10 years
        # =========================================================
        nmnths = len(pettmp_ss['precip'])
        nyrs = nmnths/12
        self.ann_ave_precip_ss = sum(pettmp_ss['precip'])/nyrs
        self.ann_ave_temp_ss = sum(pettmp_ss['tair'])/nmnths

class ReadCropOwNitrogenParms(object, ):

    def __init__(self, xls_inp_fname):
        '''
        read parameters from ORATOR inputs Excel file
        '''

        print('Reading crop, organic waste and Nitrogen parameter sheets...')

        # Nitrogen params plus r_dry, drying potential
        # ============================================
        self.n_parms = _read_n_constants_sheet(xls_inp_fname, 'N constants', 0)

        # Organic Waste and Crop params e.g. max rooting depths
        # =====================================================
        self.ow_parms = _read_organic_waste_sheet(xls_inp_fname, 'Org Waste parms', 0)
        self.crop_vars = _read_crop_vars(xls_inp_fname, 'Crop parms')

class Soil(object,):
    '''

    '''
    def __init__(self, soil_slice):
        """
        Assumptions:
        """
        self.title = 'Soil'

        t_depth = soil_slice[0]
        self.t_clay = soil_slice[1]
        self.t_silt = soil_slice[2]
        self.t_sand = soil_slice[3]
        t_carbon = soil_slice[4]
        t_bulk = soil_slice[5]
        self.t_pH_h2o = soil_slice[6]
        self.t_salinity = soil_slice[7]

        tot_soc_meas = (10 ** 4) * (t_depth / 100) * t_bulk * (t_carbon / 100)  # tonnes/ha

        self.t_depth = t_depth
        self.t_carbon = t_carbon
        self.t_bulk = t_bulk
        self.tot_soc_meas = tot_soc_meas

class Crop(object,):
    '''

    '''
    def __init__(self, crop_slice):
        """
        Assumptions:
        """
        self.title = 'Crop'

        self.crop_lu = crop_slice[0]
        self.sowing_mnth = crop_slice[1]
        self.harvest_mnth = crop_slice[2]
        self.c_yield_typ = crop_slice[3]

        # inorganic fertiliser application
        # ================================
        self.fert_type = crop_slice[5]
        self.fert_n = crop_slice[6]
        self.fert_p = crop_slice[7]
        self.fert_mnth = crop_slice[8]

        # organic waste
        # =============
        self.ow_type = crop_slice[10]
        self.ow_mnth = crop_slice[11]
        self.ow_amount = crop_slice[12]

        # irrigation up to 12 months max
        # ==============================
        indx_strt = 14
        irrig = {}
        nmnths = int(len(crop_slice[indx_strt:])/2)
        for imnth in range(nmnths):
            indx = indx_strt + imnth*2
            mnth = crop_slice[indx]
            amount = crop_slice[indx + 1]
            if amount == 0 or mnth == 0:
                continue
            else:
                irrig[mnth] = amount

        self.irrig = irrig
